#!/usr/bin/env python3
"""
Manejador de archivos para exportar datos a Excel y CSV
"""

import os
import csv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

class FileHandler:
    def __init__(self):
        self.base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        self.scrape_dir = os.path.join(self.base_dir, 'scrape')
        
        # Asegurar que existen los directorios
        self._ensure_directories()
    
    def _ensure_directories(self):
        """Crea los directorios necesarios si no existen"""
        platforms = ['instagram']
        
        for platform in platforms:
            platform_dir = os.path.join(self.scrape_dir, platform)
            os.makedirs(platform_dir, exist_ok=True)
    
    def save_to_excel(self, results, platform, filename):
        """
        Guarda los resultados en formato Excel con el formato específico requerido
        
        Args:
            results (list): Lista de resultados de scraping
            platform (str): Plataforma (instagram)
            filename (str): Nombre base del archivo
            
        Returns:
            str: Ruta completa del archivo guardado
        """
        output_dir = os.path.join(self.scrape_dir, platform)
        filepath = os.path.join(output_dir, f"{filename}.xlsx")
        
        wb = Workbook()
        ws = wb.active
        ws.title = f"{platform.upper()} Data"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill_meta = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_fill_comments = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Headers de metadatos (columnas 1-14)
        metadata_headers = [
            "Now", "Post URL", "Publisher Nickname", "Publisher @", "Publisher URL",
            "Publish Time", "Post Likes", "Post Shares", "Description",
            "Number of 1st level comments", "Number of 2nd level comments",
            "Total Comments (actual)", "Total Comments (platform says)", "Difference"
        ]
        
        # Headers de comentarios (columnas 17-28)
        comment_headers = [
            "Comment Number (ID)", "Nickname", "User @", "User URL", "Comment Text",
            "Time", "Likes", "Profile Picture URL", "Followers", "Is 2nd Level Comment",
            "User Replied To", "Number of Replies"
        ]
        
        # Añadir headers de metadatos
        for col_num, header in enumerate(metadata_headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = self._safe_unicode_value(header)
            cell.font = header_font
            cell.fill = header_fill_meta
            cell.alignment = center_alignment
            cell.border = border
        
        # Añadir headers de comentarios (empezando en columna 17)
        start_col = 17
        for col_num, header in enumerate(comment_headers):
            cell = ws.cell(row=1, column=start_col + col_num)
            cell.value = self._safe_unicode_value(header)
            cell.font = header_font
            cell.fill = header_fill_comments
            cell.alignment = center_alignment
            cell.border = border
        
        current_row = 2
        
        # Procesar cada video
        for video_result in results:
            metadata = video_result.get('metadata', {})
            comments = video_result.get('comments', [])
            
            # Añadir metadatos del video en la fila actual
            metadata_values = [
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                video_result.get('url', ''),
                metadata.get('publisher_nickname', ''),
                metadata.get('publisher_username', ''),
                metadata.get('publisher_url', ''),
                metadata.get('publish_time', ''),
                metadata.get('likes', ''),
                metadata.get('shares', ''),
                metadata.get('description', ''),
                metadata.get('level1_comments', ''),
                metadata.get('level2_comments', ''),
                len(comments),
                metadata.get('total_comments_claimed', ''),
                abs(int(metadata.get('total_comments_claimed', 0)) - len(comments)) if str(metadata.get('total_comments_claimed', '')).isdigit() else ''
            ]
            
            for col_num, value in enumerate(metadata_values, 1):
                # Ensure proper Unicode handling for emojis and special characters
                cell = ws.cell(row=current_row, column=col_num)
                cell.value = self._safe_unicode_value(value)
            
            # Añadir comentarios
            for comment_idx, comment in enumerate(comments):
                comment_values = [
                    comment_idx + 1,  # Comment Number (ID)
                    comment.get('nickname', ''),
                    comment.get('username', ''),
                    comment.get('user_url', ''),
                    comment.get('text', ''),
                    comment.get('time', ''),
                    comment.get('likes', ''),
                    comment.get('profile_pic', ''),
                    comment.get('followers', ''),
                    'Yes' if comment.get('is_reply', False) else 'No',
                    comment.get('replied_to', ''),
                    comment.get('num_replies', 0)
                ]
                
                for col_num, value in enumerate(comment_values):
                    # Ensure proper Unicode handling for emojis and special characters
                    cell = ws.cell(row=current_row, column=start_col + col_num)
                    cell.value = self._safe_unicode_value(value)
                
                if comment_idx < len(comments) - 1:  # No incrementar en el último comentario
                    current_row += 1
            
            current_row += 2  # Espacio entre videos
        
        # Ajustar ancho de columnas
        self._adjust_column_widths(ws)
        
        # Guardar archivo
        wb.save(filepath)
        wb.close()
        
        return filepath
    
    def save_to_csv(self, results, platform, filename):
        """
        Guarda los resultados en formato CSV
        
        Args:
            results (list): Lista de resultados de scraping
            platform (str): Plataforma (instagram)
            filename (str): Nombre base del archivo
            
        Returns:
            str: Ruta completa del archivo guardado
        """
        output_dir = os.path.join(self.scrape_dir, platform)
        filepath = os.path.join(output_dir, f"{filename}.csv")
        
        # Preparar datos para CSV
        all_rows = []
        
        # Header
        metadata_headers = [
            "Now", "Post URL", "Publisher Nickname", "Publisher @", "Publisher URL",
            "Publish Time", "Post Likes", "Post Shares", "Description",
            "Number of 1st level comments", "Number of 2nd level comments",
            "Total Comments (actual)", "Total Comments (platform says)", "Difference",
            "", "",  # Columnas vacías (separador)
            "Comment Number (ID)", "Nickname", "User @", "User URL", "Comment Text",
            "Time", "Likes", "Profile Picture URL", "Followers", "Is 2nd Level Comment",
            "User Replied To", "Number of Replies"
        ]
        
        all_rows.append(metadata_headers)
        
        # Procesar cada video
        for video_result in results:
            metadata = video_result.get('metadata', {})
            comments = video_result.get('comments', [])
            
            # Metadatos del video
            metadata_values = [
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                video_result.get('url', ''),
                metadata.get('publisher_nickname', ''),
                metadata.get('publisher_username', ''),
                metadata.get('publisher_url', ''),
                metadata.get('publish_time', ''),
                metadata.get('likes', ''),
                metadata.get('shares', ''),
                metadata.get('description', ''),
                metadata.get('level1_comments', ''),
                metadata.get('level2_comments', ''),
                len(comments),
                metadata.get('total_comments_claimed', ''),
                abs(int(metadata.get('total_comments_claimed', 0)) - len(comments)) if str(metadata.get('total_comments_claimed', '')).isdigit() else ''
            ]
            
            # Añadir comentarios
            for comment_idx, comment in enumerate(comments):
                row = metadata_values + ['', ''] + [  # Metadatos + separador + comentario
                    comment_idx + 1,
                    comment.get('nickname', ''),
                    comment.get('username', ''),
                    comment.get('user_url', ''),
                    comment.get('text', ''),
                    comment.get('time', ''),
                    comment.get('likes', ''),
                    comment.get('profile_pic', ''),
                    comment.get('followers', ''),
                    'Yes' if comment.get('is_reply', False) else 'No',
                    comment.get('replied_to', ''),
                    comment.get('num_replies', 0)
                ]
                
                all_rows.append(row)
                metadata_values = [''] * 14  # Vaciar metadatos para siguientes filas
            
            # Fila vacía entre videos
            all_rows.append([''] * len(metadata_headers))
        
        # Escribir CSV
        with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerows(all_rows)
        
        return filepath
    
    def _adjust_column_widths(self, worksheet):
        """Ajusta automáticamente el ancho de las columnas"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            # Ajustar ancho (mínimo 10, máximo 50)
            adjusted_width = min(max(max_length + 2, 10), 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def create_summary_report(self, results, platform, filename):
        """
        Crea un reporte resumen del scraping
        
        Args:
            results (list): Lista de resultados de scraping
            platform (str): Plataforma
            filename (str): Nombre base del archivo
            
        Returns:
            str: Ruta del archivo de reporte
        """
        output_dir = os.path.join(self.scrape_dir, platform)
        filepath = os.path.join(output_dir, f"{filename}_summary.txt")
        
        total_videos = len(results)
        total_comments = sum(len(r.get('comments', [])) for r in results)
        
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(f"REPORTE DE SCRAPING - {platform.upper()}\n")
            f.write(f"{'='*50}\n\n")
            f.write(f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Plataforma: {platform.upper()}\n")
            f.write(f"Videos procesados: {total_videos}\n")
            f.write(f"Total de comentarios extraídos: {total_comments}\n")
            f.write(f"Promedio de comentarios por video: {total_comments/total_videos:.1f}\n\n")
            
            f.write("DETALLE POR VIDEO:\n")
            f.write("-" * 30 + "\n")
            
            for i, result in enumerate(results, 1):
                url = result.get('url', 'URL no disponible')
                comments_count = len(result.get('comments', []))
                metadata = result.get('metadata', {})
                
                f.write(f"\nVideo {i}:\n")
                f.write(f"  URL: {url}\n")
                f.write(f"  Autor: {metadata.get('publisher_nickname', 'N/A')}\n")
                f.write(f"  Comentarios extraídos: {comments_count}\n")
                f.write(f"  Likes del video: {metadata.get('likes', 'N/A')}\n")
        
        return filepath
    
    def _safe_unicode_value(self, value):
        """
        Safely handle Unicode values for Excel export including emojis and accented characters
        
        Args:
            value: The value to process
            
        Returns:
            str: Properly formatted Unicode string safe for Excel
        """
        if value is None:
            return ""
        
        # Convert to string if not already
        if not isinstance(value, str):
            value = str(value)
        
        # Handle empty strings
        if not value:
            return ""
        
        try:
            # Ensure the string is properly encoded as UTF-8
            # This preserves emojis and accented characters like á, é, í, ó, ú, ñ
            if isinstance(value, bytes):
                value = value.decode('utf-8', errors='replace')
            
            # Excel handles Unicode natively with openpyxl, so we just return the string
            # This preserves emojis (🔥, 👏, 😍, etc.) and Spanish accents (á, é, í, ó, ú, ñ, ü)
            return value
            
        except (UnicodeDecodeError, UnicodeEncodeError) as e:
            # If there's any Unicode error, try to clean the string
            try:
                # Replace problematic characters but keep the basic text
                cleaned = value.encode('utf-8', errors='replace').decode('utf-8')
                return cleaned
            except:
                # Last resort: return a safe version
                return str(value).encode('ascii', errors='replace').decode('ascii')